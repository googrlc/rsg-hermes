#!/usr/bin/env python3
"""Build Zia-uploadable artifacts (csv / json / xlsx). Markdown is not accepted."""

from __future__ import annotations

import csv
import json
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "zia-upload"
DELUGE = ROOT / "deluge"
FORMS = [
    "forms_policy_master.csv",
    "forms_policy_status_history.csv",
    "forms_renewal_queue.csv",
    "forms_policy_audit.csv",
    "forms_audit_exceptions.csv",
]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str] | None = None) -> None:
    if not rows and not fieldnames:
        raise SystemExit(f"no rows for {path}")
    names = fieldnames or list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=names, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, "") for k in names})


def copy_csv(name: str) -> list[dict[str, str]]:
    rows = read_csv(ROOT / name)
    write_csv(OUT / name, rows)
    return rows


def flatten_records(records: list[dict], extra: dict | None = None) -> list[dict[str, str]]:
    rows = []
    for rec in records:
        flat = {k: "" if v is None or v is False and k != "Active" else v for k, v in rec.items()}
        if extra:
            flat.update(extra)
        rows.append({k: "" if v is None else str(v) for k, v in flat.items()})
    return rows


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    prompt = (ROOT / "ZIA_PASTE_PROMPT.md").read_text(encoding="utf-8")
    spec = (ROOT / "ZIA_AI_BUILD_INSTRUCTIONS.md").read_text(encoding="utf-8")
    readme = (ROOT / "README.md").read_text(encoding="utf-8")
    acceptance = (ROOT / "tests" / "acceptance_cases.md").read_text(encoding="utf-8")
    samples = json.loads((ROOT / "tests" / "sample_records.json").read_text(encoding="utf-8"))

    form_tables = {name: copy_csv(name) for name in FORMS}
    picklists = copy_csv("picklists.csv")
    views = copy_csv("views.csv")
    workflows = copy_csv("workflows.csv")

    all_fields = []
    for name, rows in form_tables.items():
        for row in rows:
            row = dict(row)
            row["Source_File"] = name
            all_fields.append(row)
    write_csv(OUT / "forms_all.csv", all_fields)

    deluge_rows = []
    deluge_map = {}
    for path in sorted(DELUGE.glob("*.dg")):
        code = path.read_text(encoding="utf-8")
        deluge_map[path.name] = code
        deluge_rows.append({"Install_Order": path.name[:2], "File_Name": path.name, "Deluge_Code": code})
    write_csv(OUT / "deluge_scripts.csv", deluge_rows, ["Install_Order", "File_Name", "Deluge_Code"])

    write_csv(
        OUT / "00_READ_ME_FIRST.csv",
        [
            {
                "Step": "1",
                "Action": "Upload this CSV plus RSG_Policy_Reconciliation_Zia_Pack.json and RSG_Policy_Reconciliation_Build.xlsx to Zia. Markdown cannot be uploaded.",
            },
            {"Step": "2", "Action": prompt.replace("\n", " ")},
            {
                "Step": "3",
                "Action": "Build Phase 1 only: five forms, fields from forms_*.csv, picklists.csv, status-history hook. Stop.",
            },
        ],
        ["Step", "Action"],
    )

    pm_rows = flatten_records(samples["policy_master"])
    write_csv(OUT / "sample_policy_master.csv", pm_rows)
    rq_rows = flatten_records(samples["renewal_queue"])
    write_csv(OUT / "sample_renewal_queue.csv", rq_rows)

    verdicts = [
        {"Order": "1", "Verdict": "duplicate_policy", "Severity": "High"},
        {"Order": "2", "Verdict": "pending_sync", "Severity": "Low (High if pending > 72h)"},
        {"Order": "3", "Verdict": "rewrite_detected", "Severity": "Medium"},
        {"Order": "4", "Verdict": "status_mismatch", "Severity": "Medium (High if Active vs Cancelled)"},
        {"Order": "5", "Verdict": "financial_discrepancy", "Severity": "Medium (High if |delta|>=500 or Tier A)"},
        {"Order": "6", "Verdict": "missing_in_crm", "Severity": "Medium (High if Tier A or premium>=5000)"},
        {"Order": "7", "Verdict": "missing_in_ams", "Severity": "High"},
        {"Order": "8", "Verdict": "stale_renewal_queue", "Severity": "Medium (Critical if Past Due or CRITICAL risk)"},
        {"Order": "9", "Verdict": "stale_crm", "Severity": "Low"},
        {"Order": "10", "Verdict": "cancel_reason_gap", "Severity": "Medium"},
        {"Order": "11", "Verdict": "lineage_orphan", "Severity": "Medium"},
        {"Order": "12", "Verdict": "clean_match", "Severity": "(no exception)"},
    ]
    write_csv(OUT / "verdict_matrix.csv", verdicts)

    scores = [
        {"ID": "S1", "Condition": "Policy_Number blank", "Points": "-40"},
        {"ID": "S2", "Condition": "NowCerts_Policy_GUID blank", "Points": "-25"},
        {"ID": "S3", "Condition": "CRM_Policy_ID blank", "Points": "-10"},
        {"ID": "S4", "Condition": "Insured_Name blank", "Points": "-10"},
        {"ID": "S5", "Condition": "Carrier blank", "Points": "-8"},
        {"ID": "S6", "Condition": "Line_of_Business blank", "Points": "-5"},
        {"ID": "S7", "Condition": "Effective_Date or Expiration_Date blank", "Points": "-10 each"},
        {"ID": "S8", "Condition": "Policy_Status blank or unnormalized", "Points": "-15"},
        {"ID": "S9", "Condition": "AMS vs CRM status differ", "Points": "-12"},
        {"ID": "S10", "Condition": "AMS vs Creator status differ", "Points": "-12"},
        {"ID": "S11", "Condition": "Premium blank on Active policy", "Points": "-15"},
        {"ID": "S12", "Condition": "Abs premium $ delta >= 25", "Points": "-10"},
        {"ID": "S13", "Condition": "Abs premium % delta >= 1", "Points": "-8"},
        {"ID": "S14", "Condition": "Abs premium $ delta >= 500", "Points": "-15 extra"},
        {"ID": "S15", "Condition": "Duplicate_Count > 1", "Points": "-30"},
        {"ID": "S16", "Condition": "Cancellation without class", "Points": "-12"},
        {"ID": "S17", "Condition": "Rewrite detected but Rewrite_Of empty", "Points": "-15"},
        {"ID": "S18", "Condition": "Lineage pointer set but target missing", "Points": "-12"},
        {"ID": "S19", "Condition": "Last_Synced older than 48h", "Points": "-8"},
        {"ID": "S20", "Condition": "Last_Synced older than 7 days", "Points": "-12 extra"},
        {"ID": "S21", "Condition": "Renewal_Queue stale", "Points": "-10"},
        {"ID": "S22", "Condition": "Pending_Queue_Jobs > 0", "Points": "-8"},
        {"ID": "S23", "Condition": "Active=true but status in EXCLUDE_STATUSES", "Points": "-20"},
        {"ID": "S24", "Condition": "Active=false but status = Active", "Points": "-15"},
        {"ID": "S25", "Condition": "Agency Bill Active missing Agency_Fee", "Points": "-5"},
    ]
    write_csv(OUT / "confidence_score.csv", scores)

    write_csv(
        OUT / "acceptance_cases.csv",
        [
            {
                "Seed_ID": r["seed_id"],
                "Expected_Verdict": r["expected_verdict"],
                "Expected_Confidence": r["expected_confidence"],
                "Policy_Number": r["Policy_Number"],
            }
            for r in samples["policy_master"]
        ],
    )

    pack = {
        "app_display_name": "RSG Policy Reconciliation",
        "app_link_name": "rsg_policy_reconciliation",
        "spec_version": "1.0",
        "zia_upload_formats": [".xls", ".xlsx", ".xlsm", ".csv", ".tsv", ".ods", ".mdb", ".accdb", ".ds", ".json", ".numbers"],
        "max_file_size_note": "2 GB max; files over 100 MB must be CSV. This pack is far under that.",
        "build_order": [
            "Create app rsg_policy_reconciliation",
            "Create five forms in order: Policy_Master, Policy_Status_History, Renewal_Queue, Policy_Audit, Audit_Exceptions",
            "Create fields from forms_*.csv using Deluge_Name",
            "Import picklists.csv values exactly",
            "Create lookups and related lists",
            "Install deluge_scripts.csv in Install_Order",
            "Create views.csv and workflows.csv",
            "Phase 1 stop. Then Phase 2 agent using sample_policy_master.csv",
        ],
        "hard_rules": [
            "Exactly five forms",
            "Do not invent fields or verdicts",
            "NowCerts is AMS system of record; never write NowCerts from Creator",
            "No Zoho CRM write without Approved_To_Push + Approved_By + Approved_At",
            "Never auto-create CRM Accounts",
            "Never invent policy numbers, premiums, or GUIDs",
        ],
        "zia_prompt": prompt,
        "specification_markdown": spec,
        "pack_readme": readme,
        "acceptance_markdown": acceptance,
        "forms": form_tables,
        "picklists": picklists,
        "views": views,
        "workflows": workflows,
        "verdicts": verdicts,
        "confidence_score": scores,
        "deluge": deluge_map,
        "sample_records": samples,
        "roles": {
            "Lamar": "Admin / Producer — High/Critical exceptions, CRM push approval",
            "Gretchen": "CSR — Low/Medium exceptions, cancel class, renewal touches",
        },
        "systems_of_record": {
            "NowCerts": "policy facts",
            "Zoho CRM": "Accounts, Contacts, Deals",
            "Supabase/Hermes": "ops mirror, queues, KPIs",
            "Zoho Creator": "this reconciliation workspace only",
        },
    }
    pack_path = OUT / "RSG_Policy_Reconciliation_Zia_Pack.json"
    pack_path.write_text(json.dumps(pack, indent=2), encoding="utf-8")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit("openpyxl required to write xlsx") from exc

    wb = Workbook()

    def add_sheet(title: str, rows: list[dict[str, str]], fieldnames: list[str] | None = None) -> None:
        names = fieldnames or (list(rows[0].keys()) if rows else ["(empty)"])
        ws = wb.create_sheet(title[:31])
        header_font = Font(bold=True)
        for col, name in enumerate(names, 1):
            cell = ws.cell(1, col, name)
            cell.font = header_font
        for r_i, row in enumerate(rows, 2):
            for c_i, name in enumerate(names, 1):
                val = row.get(name, "")
                if isinstance(val, str) and len(val) > 32000:
                    val = val[:31999]
                ws.cell(r_i, c_i, val)
                if name in {"Deluge_Code", "Action"}:
                    ws.cell(r_i, c_i).alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(1, len(names) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 24
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    ws0 = wb.active
    ws0.title = "ZIA_PROMPT"
    ws0["A1"] = "Paste / upload this pack into Zia. Markdown files are NOT a supported upload format."
    ws0["A1"].font = Font(bold=True)
    ws0["A3"] = prompt
    ws0["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws0.column_dimensions["A"].width = 120
    ws0.row_dimensions[3].height = 420

    add_sheet("Policy_Master", form_tables["forms_policy_master.csv"])
    add_sheet("Status_History", form_tables["forms_policy_status_history.csv"])
    add_sheet("Renewal_Queue", form_tables["forms_renewal_queue.csv"])
    add_sheet("Policy_Audit", form_tables["forms_policy_audit.csv"])
    add_sheet("Audit_Exceptions", form_tables["forms_audit_exceptions.csv"])
    add_sheet("Picklists", picklists)
    add_sheet("Views", views)
    add_sheet("Workflows", workflows)
    add_sheet("Verdicts", verdicts)
    add_sheet("Confidence_Score", scores)
    add_sheet("Deluge", deluge_rows, ["Install_Order", "File_Name", "Deluge_Code"])
    add_sheet("Sample_Policy_Master", pm_rows)
    add_sheet("Sample_Renewal_Queue", rq_rows)
    add_sheet(
        "Acceptance",
        [
            {
                "Seed_ID": r["seed_id"],
                "Expected_Verdict": r["expected_verdict"],
                "Expected_Confidence": str(r["expected_confidence"]),
                "Policy_Number": r["Policy_Number"],
            }
            for r in samples["policy_master"]
        ],
    )
    spec_sheet = wb.create_sheet("Full_Spec")
    spec_sheet["A1"] = spec
    spec_sheet["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    spec_sheet.column_dimensions["A"].width = 120
    spec_sheet.row_dimensions[1].height = 600

    xlsx_path = OUT / "RSG_Policy_Reconciliation_Build.xlsx"
    wb.save(xlsx_path)
    print(f"wrote {pack_path} ({pack_path.stat().st_size} bytes)")
    print(f"wrote {xlsx_path} ({xlsx_path.stat().st_size} bytes)")
    print(f"csv count {len(list(OUT.glob('*.csv')))}")


if __name__ == "__main__":
    main()
