#!/usr/bin/env python3
"""Render PRD/BRD/RFP PDFs and process-diagram PDF/PNGs for Zia attach."""

from __future__ import annotations

import html
import shutil
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
REQ = ROOT / "requirements"
DIA = REQ / "diagrams"
OUT = ROOT / "zia-upload"
ART = Path("/opt/cursor/artifacts")

CSS = """
@page { size: letter; margin: 0.75in; }
body { font-family: Calibri, "Segoe UI", Helvetica, Arial, sans-serif; font-size: 11pt; color: #1a1a1a; line-height: 1.4; }
h1 { font-size: 20pt; color: #0f3d5e; border-bottom: 3px solid #c45c26; padding-bottom: 6px; }
h2 { font-size: 14pt; color: #0f3d5e; margin-top: 1.2em; }
h3 { font-size: 12pt; color: #c45c26; }
p, li { orphans: 3; widows: 3; }
table { border-collapse: collapse; width: 100%; margin: 0.6em 0 1em; font-size: 10pt; }
th, td { border: 1px solid #c5d0d8; padding: 5px 7px; vertical-align: top; }
th { background: #0f3d5e; color: white; text-align: left; }
tr:nth-child(even) td { background: #f4f7f9; }
code { font-family: Consolas, monospace; font-size: 9.5pt; background: #f0f3f5; padding: 0 3px; }
.meta { color: #445; font-size: 10pt; margin-bottom: 1.2em; }
.banner { background: #0f3d5e; color: white; padding: 10px 14px; margin: -8px -8px 16px; }
.banner strong { color: #f4c79c; }
"""


def md_to_html(md: str, title: str) -> str:
    try:
        import markdown

        body = markdown.markdown(md, extensions=["tables", "fenced_code"])
    except ImportError:
        body = "<pre>" + html.escape(md) + "</pre>"
    return f"""<!doctype html><html><head><meta charset="utf-8"><title>{html.escape(title)}</title>
<style>{CSS}</style></head><body>
<div class="banner"><strong>RSG · Zoho Creator</strong> &nbsp; Attach this file to Zia as a {html.escape(title.split("—")[0].strip() if "—" in title else title)}</div>
{body}
</body></html>"""


def write_pdf(html_str: str, dest: Path) -> None:
    from weasyprint import HTML

    dest.parent.mkdir(parents=True, exist_ok=True)
    HTML(string=html_str, base_url=str(REQ)).write_pdf(dest)


def svg_box(x, y, w, h, text, fill="#0f3d5e", color="white") -> str:
    lines = text.split("\n")
    t = ""
    for i, line in enumerate(lines):
        t += f'<tspan x="{x + w/2}" dy="{"1.2em" if i else 0}">{html.escape(line)}</tspan>'
    start = y + h / 2 - (len(lines) - 1) * 7
    return f'''
    <rect x="{x}" y="{y}" width="{w}" height="{h}" rx="8" fill="{fill}"/>
    <text x="{x + w/2}" y="{start}" text-anchor="middle" fill="{color}" font-family="Calibri, Arial" font-size="13">{t}</text>'''


def svg_arrow(x1, y1, x2, y2) -> str:
    return f'<line x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}" stroke="#c45c26" stroke-width="2" marker-end="url(#arrow)"/>'


def svg_wrap(w, h, title, inner) -> str:
    return f'''<svg xmlns="http://www.w3.org/2000/svg" width="{w}" height="{h}" viewBox="0 0 {w} {h}">
  <defs>
    <marker id="arrow" markerWidth="10" markerHeight="7" refX="9" refY="3.5" orient="auto">
      <polygon points="0 0, 10 3.5, 0 7" fill="#c45c26"/>
    </marker>
  </defs>
  <rect width="100%" height="100%" fill="white"/>
  <text x="24" y="32" font-family="Calibri, Arial" font-size="18" fill="#0f3d5e" font-weight="bold">{html.escape(title)}</text>
  {inner}
</svg>'''


def diagram_systems() -> str:
    inner = (
        svg_box(40, 70, 200, 70, "NowCerts AMS\nPolicy facts SoR", "#0f3d5e")
        + svg_box(320, 70, 200, 70, "Zoho CRM\nAccounts / Deals SoR", "#1f6f8b")
        + svg_box(180, 220, 220, 70, "Zoho Creator\nRecon workspace", "#c45c26")
        + svg_box(520, 220, 200, 70, "Hermes / Supabase\nQueues · KPIs · P85", "#4a5560")
        + svg_arrow(140, 140, 250, 220)
        + svg_arrow(420, 140, 320, 220)
        + svg_arrow(510, 255, 400, 255)
        + '<text x="40" y="360" font-family="Calibri, Arial" font-size="12" fill="#333">Creator copies AMS facts, stamps CRM IDs, never writes NowCerts, writes CRM only with approval.</text>'
    )
    return svg_wrap(760, 390, "1. Systems of record", inner)


def diagram_daily() -> str:
    boxes = [
        (40, 70, "06:40 ET\nCRM pull (Ph4)"),
        (220, 70, "06:45 ET\nSelect A/B/window"),
        (400, 70, "Normalize\nstatus/billing"),
        (580, 70, "Detect rewrite"),
        (40, 200, "Score 0–100"),
        (220, 200, "Verdict\nfirst-match"),
        (400, 200, "Insert\nPolicy_Audit"),
        (580, 200, "Upsert\nexception + SLA"),
        (310, 320, "Stamp Policy_Master\nLast_Verdict / Confidence"),
    ]
    inner = ""
    for x, y, t in boxes:
        inner += svg_box(x, y, 150, 70, t)
    inner += svg_arrow(190, 105, 220, 105)
    inner += svg_arrow(370, 105, 400, 105)
    inner += svg_arrow(550, 105, 580, 105)
    inner += svg_arrow(655, 140, 655, 200)
    inner += svg_arrow(580, 235, 550, 235)
    inner += svg_arrow(400, 235, 370, 235)
    inner += svg_arrow(220, 235, 190, 235)
    inner += svg_arrow(115, 270, 115, 355)
    inner += svg_arrow(115, 355, 310, 355)
    return svg_wrap(780, 430, "2. Daily reconciliation (one Run_ID)", inner)


def diagram_verdict() -> str:
    items = [
        "1 duplicate_policy",
        "2 pending_sync",
        "3 rewrite_detected",
        "4 status_mismatch",
        "5 financial_discrepancy",
        "6 missing_in_crm",
        "7 missing_in_ams",
        "8 stale_renewal_queue",
        "9 stale_crm",
        "10 cancel_reason_gap",
        "11 lineage_orphan",
        "12 clean_match",
    ]
    inner = '<text x="24" y="58" font-family="Calibri, Arial" font-size="12" fill="#333">First match wins. Do not add or rename types.</text>'
    for i, label in enumerate(items):
        col = i % 3
        row = i // 3
        x = 40 + col * 240
        y = 80 + row * 70
        fill = "#2f855a" if "clean_match" in label else "#0f3d5e"
        inner += svg_box(x, y, 220, 54, label, fill)
        if i < 11 and col < 2:
            inner += svg_arrow(x + 220, y + 27, x + 240, y + 27)
    return svg_wrap(760, 380, "3. Verdict decision tree (order)", inner)


def diagram_cancel() -> str:
    inner = (
        svg_box(40, 70, 180, 64, "Status in cancel set", "#0f3d5e")
        + svg_box(280, 70, 200, 64, "Cancellation_Class\nrequired", "#c45c26")
        + svg_box(540, 40, 180, 50, "Non Pay\nreinstatement", "#4a5560")
        + svg_box(540, 100, 180, 50, "Insured Request\nreplacement task", "#4a5560")
        + svg_box(540, 160, 180, 50, "Underwriter\nLamar if commercial", "#4a5560")
        + svg_box(280, 200, 200, 64, "Rewrite?\nsame insured+LOB\n±60 days", "#1f6f8b")
        + svg_box(40, 300, 180, 64, "One successor\nlink Rewrite_Of", "#2f855a")
        + svg_box(280, 300, 200, 64, "Many successors\nHigh exception", "#9b2c2c")
        + svg_box(540, 300, 180, 64, "None found\nstill class Rewrite\nif human said so", "#4a5560")
        + svg_arrow(220, 102, 280, 102)
        + svg_arrow(480, 90, 540, 65)
        + svg_arrow(380, 134, 380, 200)
        + svg_arrow(380, 264, 130, 300)
        + svg_arrow(380, 264, 380, 300)
        + svg_arrow(380, 264, 630, 300)
    )
    return svg_wrap(760, 400, "4. Cancellation and rewrite", inner)


def diagram_exception() -> str:
    inner = (
        svg_box(40, 80, 200, 70, "Verdict ≠ clean_match\nor score < 80\nor $/% delta", "#0f3d5e")
        + svg_box(300, 80, 200, 70, "Severity matrix\n+ bump if score<50", "#1f6f8b")
        + svg_box(560, 80, 180, 70, "SLA hours\n4 / 24 / 72 / 120", "#c45c26")
        + svg_box(40, 220, 200, 70, "Upsert same\npolicy+verdict", "#0f3d5e")
        + svg_box(300, 220, 200, 70, "Owner: Low/Med Gretchen\nHigh/Crit Lamar", "#1f6f8b")
        + svg_box(560, 220, 180, 70, "Hourly SLA sweep\nnotify on breach", "#c45c26")
        + svg_box(300, 340, 200, 64, "Auto-close only if\nwas pending_sync\nor stale_crm", "#2f855a")
        + svg_arrow(240, 115, 300, 115)
        + svg_arrow(500, 115, 560, 115)
        + svg_arrow(140, 150, 140, 220)
        + svg_arrow(240, 255, 300, 255)
        + svg_arrow(500, 255, 560, 255)
        + svg_arrow(400, 290, 400, 340)
    )
    return svg_wrap(780, 430, "5. Exception and SLA", inner)


def diagram_crm_write() -> str:
    inner = (
        svg_box(40, 80, 220, 80, "Policy_Audit\nRecommended_Payload", "#0f3d5e")
        + svg_box(320, 80, 220, 80, "Lamar sets\nApproved_To_Push\n+ By + At", "#c45c26")
        + svg_box(600, 80, 140, 80, "At < 24h?", "#1f6f8b")
        + svg_box(320, 220, 220, 80, "zoho.crm.updateRecord\nPolicies only", "#0f3d5e")
        + svg_box(40, 220, 220, 80, "STOP\ngate failed", "#9b2c2c")
        + svg_box(600, 220, 140, 80, "Clear flag\nSync=Pending\nhistory row", "#2f855a")
        + '<text x="40" y="360" font-family="Calibri, Arial" font-size="12" fill="#333">Never create Accounts. Never call NowCerts from Creator. Policy create drafts stay human.</text>'
        + svg_arrow(260, 120, 320, 120)
        + svg_arrow(540, 120, 600, 120)
        + svg_arrow(670, 160, 670, 220)
        + svg_arrow(600, 160, 150, 220)
        + svg_arrow(540, 260, 600, 260)
    )
    return svg_wrap(780, 390, "6. Approval-gated CRM write", inner)


def png_from_svg(svg: str, dest: Path) -> None:
    """Rasterize SVG via WeasyPrint HTML wrapper."""
    wrap = f"""<!doctype html><html><head><meta charset="utf-8">
    <style>@page {{ size: 820px 460px; margin: 0; }} body {{ margin: 0; }}</style>
    </head><body>{svg}</body></html>"""
    from weasyprint import HTML

    # PDF page then we still want PNG. Use cairosvg if present else PDF only.
    try:
        import cairosvg

        cairosvg.svg2png(bytestring=svg.encode("utf-8"), write_to=str(dest), output_width=1600)
    except Exception:
        HTML(string=wrap).write_pdf(dest.with_suffix(".pdf"))


def main() -> None:
    REQ.mkdir(parents=True, exist_ok=True)
    DIA.mkdir(parents=True, exist_ok=True)
    OUT.mkdir(parents=True, exist_ok=True)
    ART.mkdir(parents=True, exist_ok=True)

    docs = [
        ("PRD_RSG_Policy_Reconciliation.md", "PRD — Product Requirements", "PRD_RSG_Policy_Reconciliation.pdf"),
        ("BRD_RSG_Policy_Reconciliation.md", "BRD — Business Requirements", "BRD_RSG_Policy_Reconciliation.pdf"),
        ("RFP_RSG_Policy_Reconciliation.md", "RFP — Statement of Work for Zia", "RFP_RSG_Policy_Reconciliation.pdf"),
    ]
    pdfs = []
    for md_name, title, pdf_name in docs:
        md = (REQ / md_name).read_text(encoding="utf-8")
        pdf_path = REQ / pdf_name
        write_pdf(md_to_html(md, title), pdf_path)
        pdfs.append(pdf_path)
        print("pdf", pdf_path, pdf_path.stat().st_size)

    diagrams = [
        ("01_systems_of_record.svg", diagram_systems()),
        ("02_daily_reconciliation.svg", diagram_daily()),
        ("03_verdict_tree.svg", diagram_verdict()),
        ("04_cancellation_rewrite.svg", diagram_cancel()),
        ("05_exception_sla.svg", diagram_exception()),
        ("06_crm_write_gate.svg", diagram_crm_write()),
    ]
    for name, svg in diagrams:
        p = DIA / name
        p.write_text(svg, encoding="utf-8")
        png_from_svg(svg, DIA / name.replace(".svg", ".png"))

    fig_html = ["<!doctype html><html><head><meta charset='utf-8'><title>Process Diagrams</title>"]
    fig_html.append(f"<style>{CSS} svg {{ width: 100%; height: auto; margin: 12px 0 28px; border: 1px solid #e2e8f0; }}</style></head><body>")
    fig_html.append('<div class="banner"><strong>RSG · Process diagrams</strong> &nbsp; Attach this PDF to Zia with the PRD, BRD, and RFP</div>')
    fig_html.append("<h1>RSG Policy Reconciliation — Process Diagrams</h1>")
    fig_html.append("<p>Six required flows. Implement these in Deluge; do not substitute.</p>")
    for name, svg in diagrams:
        fig_html.append(svg)
    fig_html.append("</body></html>")
    diagrams_pdf = REQ / "PROCESS_DIAGRAMS_RSG_Policy_Reconciliation.pdf"
    write_pdf("\n".join(fig_html), diagrams_pdf)
    pdfs.append(diagrams_pdf)
    print("pdf", diagrams_pdf, diagrams_pdf.stat().st_size)

    # Combined package PDF (PRD+BRD+RFP+diagrams) via weasyprint concatenation is awkward;
    # copy individuals into zia-upload and artifacts.
    for p in pdfs:
        shutil.copyfile(p, OUT / p.name)
        shutil.copyfile(p, ART / p.name)
        shutil.copyfile(p, ROOT / p.name)

    for svg_path in DIA.glob("*.svg"):
        shutil.copyfile(svg_path, ART / svg_path.name)
    for png_path in DIA.glob("*.png"):
        shutil.copyfile(png_path, ART / png_path.name)

    # Spreadsheet fallback: PRD/BRD/RFP as text sheets (Zia data-upload dialog).
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "HOW_TO_ATTACH"
    ws["A1"] = "Attach these requirement files to Zia: PRD PDF, BRD PDF, RFP PDF, PROCESS_DIAGRAMS PDF. If the dialog only accepts spreadsheets, upload this XLSX plus ZIA_UPLOAD.xlsx."
    ws["A1"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 120
    for md_name, title, _pdf in docs:
        sheet = wb.create_sheet(title.split("—")[0].strip()[:31])
        text = (REQ / md_name).read_text(encoding="utf-8")
        sheet["A1"] = text
        sheet["A1"].alignment = Alignment(wrap_text=True, vertical="top")
        sheet["A1"].font = Font(name="Calibri", size=11)
        sheet.column_dimensions["A"].width = 120
        sheet.row_dimensions[1].height = 600
    xlsx = ROOT / "ZIA_REQUIREMENTS.xlsx"
    wb.save(xlsx)
    shutil.copyfile(xlsx, OUT / "ZIA_REQUIREMENTS.xlsx")
    shutil.copyfile(xlsx, ART / "ZIA_REQUIREMENTS.xlsx")
    print("xlsx", xlsx, xlsx.stat().st_size)


if __name__ == "__main__":
    main()
