"""Deterministic document extraction — file -> SubmissionObject fields.

Phase 1 is honest and testable: classify by filename/extension, pull text
(pypdf for PDFs, openpyxl for xlsx), then a regex/keyword pass maps text onto
spine fields. **XDATE gets the most thorough treatment.** Anything not
confidently extracted stays ``None`` -> validators flag it -> the human fixes it
in the review UI. The gate catches what extraction misses; that's the design,
not a failure. LLM-assisted extraction is a later upgrade behind this interface.

Never fabricate: a value we don't read stays absent, and everything we do set
records provenance in ``enrichment.sources``.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from .submission import SubmissionObject

# filename keyword -> doc_type (first hit wins)
_DOC_HINTS: list[tuple[str, str]] = [
    ("dec", "dec_page"), ("declaration", "dec_page"),
    ("mvr", "mvr"), ("motor vehicle", "mvr"),
    ("license", "drivers_license"), ("_dl", "drivers_license"),
    ("prior", "prior_policy"), ("policy", "prior_policy"),
    ("payroll", "payroll"), ("census", "census"), ("sbc", "sbc"),
    ("recording", "recording"), (".mp3", "recording"), (".wav", "recording"), (".m4a", "recording"),
]


def classify_doc(filename: str) -> str:
    name = (filename or "").lower()
    for hint, doc_type in _DOC_HINTS:
        if hint in name:
            return doc_type
    return "other"


def read_text(path: str | Path) -> str:
    """Best-effort text from a PDF/xlsx/txt. Returns '' when there's no text
    layer (scanned PDF) — that's the signal a later OCR tier is needed."""
    p = Path(path)
    ext = p.suffix.lower()
    if ext == ".pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(p))
            return "\n".join((page.extract_text() or "") for page in reader.pages)
        except Exception:
            return ""
    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(p), read_only=True, data_only=True)
            out: list[str] = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    out.append(" ".join("" if c is None else str(c) for c in row))
            return "\n".join(out)
        except Exception:
            return ""
    if ext in (".txt", ".md", ".csv"):
        try:
            return p.read_text(encoding="utf-8", errors="replace")
        except Exception:
            return ""
    return ""


# Below this many characters we treat a PDF's text layer as empty/thin (a
# scanned or image-only page) and hand it to the OCR tier.
_THIN_TEXT_CHARS = 40


def read_document_text(path: str | Path, *, ocr: bool = True) -> str:
    """OCR-aware text read. Tries the deterministic text layer first; if a PDF
    comes back empty/thin (scanned page), falls back to the vision OCR tier.

    This is the reader intake and the general extractor use so image-only quotes
    and dec pages stop coming back blank. ``read_text`` stays the pure text-layer
    path for callers that must not spend an LLM call."""
    text = read_text(path)
    if not ocr or Path(path).suffix.lower() != ".pdf":
        return text
    if len(text.strip()) >= _THIN_TEXT_CHARS:
        return text
    from .ocr import ocr_pdf

    return ocr_pdf(path) or text


# ---- field extraction ----------------------------------------------------

_DATE_RX = re.compile(
    r"(\d{1,2}/\d{1,2}/\d{2,4}|\d{4}-\d{2}-\d{2}|[A-Za-z]{3,9}\s+\d{1,2},?\s+\d{4})"
)
_XDATE_LABEL = re.compile(
    r"(expiration\s*date|exp\.?\s*date|expires?|renewal\s*date|x[-\s]?date)", re.I
)
_MONEY_RX = re.compile(r"\$?\s*([\d,]+(?:\.\d{2})?)")


def _parse_date(s: str) -> Optional[date]:
    s = s.strip().replace(",", "")
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%B %d %Y", "%b %d %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def extract_xdate(text: str) -> Optional[date]:
    """Find the expiring policy's expiration date — the XDATE. Looks for a date
    immediately after an expiration/renewal/x-date label (the reliable signal),
    so an unrelated effective date doesn't get mistaken for it."""
    for m in _XDATE_LABEL.finditer(text or ""):
        window = text[m.end(): m.end() + 40]
        dm = _DATE_RX.search(window)
        if dm:
            d = _parse_date(dm.group(1))
            if d:
                return d
    return None


def _labeled(text: str, label: str, width: int = 60) -> Optional[str]:
    m = re.search(label + r"\s*[:\-]?\s*(.+)", text or "", re.I)
    if not m:
        return None
    return m.group(1).strip()[:width].strip() or None


def extract_fields(text: str, doc_type: str = "dec_page") -> dict[str, Any]:
    """Text -> {spine_path: value}. Conservative: only confident hits."""
    fields: dict[str, Any] = {}

    xd = extract_xdate(text)
    if xd:
        fields["current_policy_expiration"] = xd

    carrier = _labeled(text, r"(?:carrier|insurer|company)")
    if carrier:
        # drop trailing noise like "Policy Number: ..."
        carrier = re.split(r"\s{2,}|policy|number", carrier, flags=re.I)[0].strip()
        if carrier:
            fields["current_carrier"] = carrier

    prem_line = _labeled(text, r"(?:total\s*premium|premium)")
    if prem_line:
        mm = _MONEY_RX.search(prem_line)
        if mm:
            try:
                fields["current_premium"] = float(mm.group(1).replace(",", ""))
            except ValueError:
                pass

    name = _labeled(text, r"(?:named\s*insured|insured\s*name|insured)")
    if name:
        name = re.split(r"\s{2,}", name)[0].strip()
        if name:
            fields["client_name"] = name

    return fields


def apply_extraction(sub: SubmissionObject, fields: dict[str, Any], source: str) -> SubmissionObject:
    """Gap-fill the submission from extracted fields (never overwrite a value
    the submitter already gave) and record provenance per field."""
    for path, value in fields.items():
        if value in (None, ""):
            continue
        if getattr(sub, path, None) in (None, ""):
            setattr(sub, path, value)
            sub.enrichment.sources[path] = source
    return sub
