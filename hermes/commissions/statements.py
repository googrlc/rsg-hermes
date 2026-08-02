"""Commission statement upload — parse, stage, review, commit.

Nothing about this path writes money without a human saying so. The stages are
separate tables on purpose, so an approval is a decision about something already
parsed and checked rather than a leap of faith about a file:

    upload  -> commission_ingest_batches   (content_hash, crosscheck, flags)
            -> commission_transactions_staging  (parsed lines, raw_row kept)
            -> review card                 (matched / created / unmatched / negatives)
    APPROVE -> commission_statements       (the header)
            -> commission_transactions     (the committed lines)
            -> matching ladder + rollup    (actual/status recomputed)

DEDUPE is enforced by the database, not by us: commission_ingest_batches has
UNIQUE (content_hash). Re-uploading the same file is rejected before a single
line is parsed, which is the only way to be sure a statement cannot be
double-counted.

THE CROSSCHECK is the other guard. Carriers print their own totals on the
statement. If what we parsed doesn't add up to what the carrier says it paid,
the parse is wrong and committing it would put fiction in the ledger. A batch
whose crosscheck fails cannot be approved.

Amounts are Decimal end to end. Statement money parsed through float is how you
get a $0.01 discrepancy on every row and a reconciliation surface nobody trusts.
"""

from __future__ import annotations

import csv
import hashlib
import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from hermes_integrations.supabase_client import SupabaseClient

log = logging.getLogger(__name__)

BATCHES_TABLE = "commission_ingest_batches"
STAGING_TABLE = "commission_transactions_staging"
STATEMENTS_TABLE = "commission_statements"
TRANSACTIONS_TABLE = "commission_transactions"

STATUS_PENDING_REVIEW = "pending_review"
STATUS_APPROVED = "approved"
STATUS_REJECTED = "rejected"
STATUS_COMMITTED = "committed"
STATUS_NEEDS_MAPPING = "needs_mapping"
STATUS_ERROR = "error"

# How a file was read. Mirrors commission_ingest_batches.extraction_method.
METHOD_CSV = "csv"
METHOD_XLSX = "xlsx"
METHOD_PDF_TEXT = "pdf_text"
METHOD_PDF_OCR = "pdf_ocr"

# Reading methods that carry column-alignment risk a human has to rule out.
# A CSV names its own columns; a PDF's columns are inferred from geometry (text
# tier) or from a picture (OCR tier), and a column read one position off looks
# exactly like a clean parse. So a PDF batch always needs someone to say they
# checked the numbers against the document.
CONFIRM_REQUIRED_METHODS = frozenset({METHOD_PDF_TEXT, METHOD_PDF_OCR})


def requires_source_confirmation(method: Any, *, is_ocr: Any = False) -> bool:
    """Does a batch read this way need an explicit human attestation to commit?

    Matched on the ``pdf`` PREFIX rather than against the exact set, because the
    set is not the whole truth about what is in the table: the Slack-drop poller
    that predates this code wrote ``extraction_method='pdf'``, and two such rows
    are on file right now. An exact-match gate would wave those straight through
    the one check they most need — the whole point is that nobody can read a
    column off a PDF and be sure, whatever the row calls the method.

    ``is_ocr`` is honoured independently: a batch flagged as machine-read needs
    confirming even if its method string says something else entirely.
    """
    text = str(method or "").strip().lower()
    return text.startswith("pdf") or bool(is_ocr)

# Columns Postgres computes. Sending any of them raises 428C9 "cannot insert a
# non-DEFAULT value". Listed per table so a new one is a one-line change rather
# than another live failure — this codebase has now been bitten twice, by
# commission_ledger.delta and commission_transactions.is_negative.
GENERATED_COLUMNS: dict[str, frozenset[str]] = {
    "commission_transactions": frozenset({"is_negative"}),
    "commission_ledger": frozenset({"delta", "unearned_balance"}),
    "commission_audits": frozenset({"variance"}),
    "commission_rules": frozenset({"lookup_priority"}),
}


def strip_generated(table: str, payload: dict[str, Any]) -> dict[str, Any]:
    """Drop columns the database computes. Postgres owns them; we never send them."""
    generated = GENERATED_COLUMNS.get(table, frozenset())
    return {k: v for k, v in payload.items() if k not in generated}

# How far parsed totals may drift from the carrier's own stated totals before we
# refuse to commit. A cent or two is rounding; anything more means the parse
# missed rows or double-counted them.
CROSSCHECK_TOLERANCE = Decimal("1.00")

# Column aliases seen across carrier statements. Matched case-insensitively
# after stripping non-alphanumerics, so "Policy #", "policy_number" and
# "POLICY NO." all land on the same field.
_ALIASES: dict[str, tuple[str, ...]] = {
    "policy_number": ("policynumber", "policy", "policyno", "policynum", "polnum"),
    "insured_name": ("insuredname", "insured", "client", "clientname", "namedinsured"),
    "carrier_name": ("carriername", "carrier", "company", "companyname"),
    "lob": ("lob", "lineofbusiness", "prod", "product", "coverage"),
    "transaction_code": ("trancode", "transactioncode", "type", "transactiontype"),
    "transaction_date": ("trandate", "transactiondate", "date", "effectivedate", "paiddate"),
    "month_key": ("monthend", "monthkey", "statementmonth", "period"),
    "gross_premium": ("grosspremium", "premium", "writtenpremium"),
    "commission_rate": ("commrate", "commissionrate", "rate"),
    "commission_amount": ("grosscomm", "commissionamount", "commission", "netdueagent",
                          "commissionpaid", "amount", "amountpaid"),
    "fee_amount": ("feeamount", "fee", "fees"),
    "agency_due": ("agencydue",),
}

# transaction_code -> the normalized type the ledger reasons about.
#
# This vocabulary is not a fresh choice — it is the one already in the data. The
# 182 rows on file distinguish renewal (57), endorsement (43), adjustment (29),
# cancel (18), new (17), fee (13) and reinstatement (5), written by the
# browser-side parsers. Collapsing endorsement/cancel/reinstatement into
# "adjustment" here, while this path becomes the only writer, would erase a
# distinction the history carries and leave the same event named two ways
# depending on which year it was loaded.
#
# Matched as substrings, first rule wins, so the order is load-bearing:
#   * "credit endorsement" is a premium credit, NOT a policy endorsement, and
#     must be tested before the bare "endorsement" rule.
#   * "new" must come last — it is a substring of "renewal".
_TYPE_RULES: tuple[tuple[str, str], ...] = (
    ("new business", "new"),
    ("newbusiness", "new"),
    ("renewal", "renewal"),
    ("reinstate", "reinstatement"),
    ("chargeback", "chargeback"),
    ("credit endorsement", "adjustment"),
    ("endorsement", "endorsement"),
    ("cancel", "cancel"),
    ("credit", "adjustment"),
    ("audit", "adjustment"),
    ("adjust", "adjustment"),
    ("fee", "fee"),
    ("new", "new"),
)


def _key(name: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(name or "").lower())


def content_hash(content: bytes) -> str:
    """SHA-256 of the raw file. The dedupe key, enforced UNIQUE by the DB."""
    return hashlib.sha256(content).hexdigest()


def as_money(value: Any) -> Decimal | None:
    """Parse statement money. Handles $, commas, and (123.45) for negatives."""
    if value is None or value == "":
        return None
    text = str(value).strip().replace("$", "").replace(",", "")
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    if not text or text in {"-", "."}:
        return None
    try:
        amount = Decimal(text)
    except InvalidOperation:
        return None
    return -amount if negative else amount


def as_rate(value: Any) -> Decimal | None:
    """Parse a commission rate. 15%, 0.15 and 15 all mean 15%.

    A bare number above 1 is read as a percentage — no carrier pays 1500%, and
    reading 15 as 1500% would corrupt every expected-commission comparison.
    """
    if value is None or value == "":
        return None
    text = str(value).strip()
    percent = text.endswith("%")
    try:
        amount = Decimal(text.rstrip("%").replace(",", ""))
    except InvalidOperation:
        return None
    if percent or amount > 1:
        return amount / Decimal("100")
    return amount


_DATE_FORMATS = ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%Y%m%d", "%m-%d-%Y")


def as_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def as_month_key(value: Any, fallback: date | None = None) -> int | None:
    """YYYYMM. Carriers write it as 202602, 2026-02, or not at all."""
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) >= 6:
        try:
            candidate = int(digits[:6])
            if 190001 <= candidate <= 299912:
                return candidate
        except ValueError:
            pass
    if fallback:
        return fallback.year * 100 + fallback.month
    return None


def normalize_transaction_type(code: Any) -> str:
    text = str(code or "").strip().lower()
    for needle, kind in _TYPE_RULES:
        if needle in text:
            return kind
    return "other"


def _field(row: dict[str, Any], name: str) -> Any:
    """Pull a logical field from a raw row via the alias table."""
    wanted = (_key(name), *_ALIASES.get(name, ()))
    for raw_key, value in row.items():
        if _key(raw_key) in wanted and value not in ("", None):
            return value
    return None


def parse_row(raw: dict[str, Any]) -> dict[str, Any] | None:
    """One statement row -> a staging line. ``None`` if it isn't a real line.

    The whole original row is kept in ``raw_row``: when a carrier disputes a
    commission, the line as they sent it is the evidence.
    """
    policy_number = str(_field(raw, "policy_number") or "").strip()
    amount = as_money(_field(raw, "commission_amount"))
    if not policy_number and amount is None:
        return None                       # blank or a layout artefact

    when = as_date(_field(raw, "transaction_date"))
    code = _field(raw, "transaction_code")

    return {
        "policy_number": policy_number,
        "insured_name": str(_field(raw, "insured_name") or "").strip() or None,
        "carrier_name": str(_field(raw, "carrier_name") or "").strip() or None,
        "lob": str(_field(raw, "lob") or "").strip() or None,
        "transaction_code": str(code or "").strip() or None,
        "transaction_type": normalize_transaction_type(code),
        "transaction_date": when.isoformat() if when else None,
        "month_key": as_month_key(_field(raw, "month_key"), when),
        "gross_premium": as_money(_field(raw, "gross_premium")),
        "commission_rate": as_rate(_field(raw, "commission_rate")),
        "commission_amount": amount,
        "fee_amount": as_money(_field(raw, "fee_amount")),
        "raw_row": {k: (str(v) if v is not None else None) for k, v in raw.items()},
    }


def _policyless_warning(rows: list[str]) -> list[str]:
    """Rows carrying money but no policy number are EXCLUDED from the parse.

    Almost always a subtotal or grand-total line. Including one double-counts
    the statement — a generated 3-line test workbook with a TOTAL row parsed as
    $289.18 against a true $144.59. And a line with no policy number could never
    match a ledger row anyway, so it has nowhere to go.

    Reported, never silent: if one of these really is a payment, a human needs to
    see it and fix the statement.
    """
    if not rows:
        return []
    shown = ", ".join(rows[:5])
    more = f" (+{len(rows) - 5} more)" if len(rows) > 5 else ""
    return [
        f"{len(rows)} row(s) carried a commission amount with NO policy number and "
        f"were excluded as subtotal/total lines: {shown}{more}"
    ]


def parse_csv(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """Parse a CSV/TSV statement. Returns (lines, warnings)."""
    warnings: list[str] = []
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
        warnings.append("file was not UTF-8; decoded as latin-1")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    lines: list[dict[str, Any]] = []
    skipped = 0
    policyless: list[str] = []
    for index, raw in enumerate(csv.DictReader(io.StringIO(text), dialect=dialect), start=2):
        parsed = parse_row(raw)
        if parsed is None:
            skipped += 1
            continue
        if not parsed["policy_number"]:
            policyless.append(f"row {index} ({parsed['commission_amount']})")
            continue
        lines.append(parsed)
    if skipped:
        warnings.append(f"{skipped} row(s) skipped as blank or non-data")
    warnings.extend(_policyless_warning(policyless))
    return lines, warnings


def parse_xlsx(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """Parse an .xlsx statement via openpyxl.

    Carriers put the header on whatever row they like, under a title block or a
    merged logo. So the header is *found* rather than assumed: the first row
    whose cells look like the columns we need. Assuming row 1 would silently
    parse a title banner as column names and yield zero lines.
    """
    warnings: list[str] = []
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - dependency is declared
        return [], ["openpyxl is not installed; cannot parse .xlsx"]

    try:
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return [], [f"could not open workbook: {exc}"]

    sheet = book.active
    if sheet is None:
        return [], ["workbook has no active sheet"]

    sheet_count = len(book.sheetnames)
    sheet_title = sheet.title
    rows = [
        ["" if c is None else c for c in row]
        for row in sheet.iter_rows(values_only=True)
    ]
    book.close()
    if not rows:
        return [], ["worksheet is empty"]

    # Find the header: the first row containing a recognisable policy column AND
    # a recognisable amount column.
    wanted_policy = {_key("policy_number"), *_ALIASES["policy_number"]}
    wanted_amount = {_key("commission_amount"), *_ALIASES["commission_amount"]}
    header_at = None
    for index, row in enumerate(rows[:25]):
        keys = {_key(cell) for cell in row}
        if keys & wanted_policy and keys & wanted_amount:
            header_at = index
            break
    if header_at is None:
        return [], ["could not find a header row with a policy and a commission column"]
    if header_at:
        warnings.append(f"header found on row {header_at + 1}, not row 1")

    header = [str(cell).strip() for cell in rows[header_at]]
    lines: list[dict[str, Any]] = []
    skipped = 0
    policyless: list[str] = []
    for offset, row in enumerate(rows[header_at + 1:], start=header_at + 2):
        raw = {header[i]: row[i] for i in range(min(len(header), len(row))) if header[i]}
        parsed = parse_row(raw)
        if parsed is None:
            skipped += 1
            continue
        if not parsed["policy_number"]:
            policyless.append(f"row {offset} ({parsed['commission_amount']})")
            continue
        lines.append(parsed)
    if skipped:
        warnings.append(f"{skipped} row(s) skipped as blank or non-data")
    warnings.extend(_policyless_warning(policyless))
    if sheet_count > 1:
        warnings.append(
            f"workbook has {sheet_count} sheets; only '{sheet_title}' was parsed"
        )
    return lines, warnings


def parse_pdf(content: bytes) -> tuple[list[dict[str, Any]], list[str], str]:
    """Parse a statement PDF. Returns (lines, warnings, extraction_method).

    The reading itself lives in ``pdf.py`` — text layer first, vision OCR only
    when there is no text layer to read. What comes back is raw header->cell
    dicts, so the rows land in ``parse_row`` exactly like a CSV's would and the
    subtotal-line exclusion applies unchanged.
    """
    from hermes.commissions.pdf import read_pdf

    extraction = read_pdf(content)
    warnings = list(extraction.warnings)

    lines: list[dict[str, Any]] = []
    skipped = 0
    policyless: list[str] = []
    for index, raw in enumerate(extraction.rows, start=1):
        parsed = parse_row(raw)
        if parsed is None:
            skipped += 1
            continue
        if not parsed["policy_number"]:
            policyless.append(f"row {index} ({parsed['commission_amount']})")
            continue
        lines.append(parsed)
    if skipped:
        warnings.append(f"{skipped} row(s) skipped as blank or non-data")
    warnings.extend(_policyless_warning(policyless))
    return lines, warnings, extraction.method


@dataclass
class ParsedFile:
    """A statement file read into lines, and everything the file told us.

    ``carrier`` and the stated totals are populated only by a carrier-specific
    parser, which knows where its statement prints them. The generic reader
    leaves them None and the uploader supplies them.
    """

    lines: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    method: str = METHOD_CSV
    parser_key: str = "generic_v1"
    carrier: str | None = None
    stated_premium: Decimal | None = None
    stated_commission: Decimal | None = None


def parse_statement(content: bytes, filename: str) -> tuple[list[dict[str, Any]], list[str]]:
    """Dispatch on extension. See ``parse_file`` for everything else the file says."""
    parsed = parse_file(content, filename)
    return parsed.lines, parsed.warnings


def parse_upload(
    content: bytes, filename: str,
) -> tuple[list[dict[str, Any]], list[str], str]:
    """Lines, warnings and HOW the file was read."""
    parsed = parse_file(content, filename)
    return parsed.lines, parsed.warnings, parsed.method


def parse_file(content: bytes, filename: str) -> ParsedFile:
    """Read a statement: carrier-specific parser first, generic reader second.

    A carrier parser is preferred wherever one recognises the file, because it
    carries knowledge the alias table cannot express — which column is the
    incremental one, which line is a fee rather than a commission, and where the
    carrier prints its own totals. The generic reader handles everyone else.

    The extraction *method* is about the container (csv/xlsx/pdf); the *parser
    key* is about who wrote the file. Both land on the batch.
    """
    from hermes.commissions.carriers import parse_carrier

    suffix = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()

    if suffix in {"csv", "tsv", "txt", "xlsx", "xlsm"}:
        method = METHOD_XLSX if suffix in {"xlsx", "xlsm"} else METHOD_CSV
        carrier_parse = parse_carrier(content, filename)
        if carrier_parse is not None and carrier_parse.lines:
            return ParsedFile(
                lines=carrier_parse.lines, warnings=carrier_parse.warnings,
                method=method, parser_key=carrier_parse.parser_key,
                carrier=carrier_parse.carrier,
                stated_premium=carrier_parse.stated_premium,
                stated_commission=carrier_parse.stated_commission,
            )
        lines, warnings = parse_csv(content) if method == METHOD_CSV else parse_xlsx(content)
        return ParsedFile(lines=lines, warnings=warnings, method=method,
                          parser_key=f"{method}_generic_v1")

    if suffix == "xls":
        return ParsedFile(warnings=["legacy .xls is not supported — re-save as .xlsx or CSV"])
    if suffix == "pdf":
        lines, warnings, method = parse_pdf(content)
        return ParsedFile(lines=lines, warnings=warnings, method=method,
                          parser_key=f"{method}_generic_v1")
    return ParsedFile(warnings=[f"unsupported statement format: .{suffix or 'unknown'}"])


@dataclass
class Crosscheck:
    """Parsed totals against what the carrier says it paid."""

    parsed_premium: Decimal = Decimal("0")
    parsed_commission: Decimal = Decimal("0")
    stated_premium: Decimal | None = None
    stated_commission: Decimal | None = None

    @property
    def commission_delta(self) -> Decimal | None:
        if self.stated_commission is None:
            return None
        return self.parsed_commission - self.stated_commission

    @property
    def ok(self) -> bool:
        """True when we can prove the parse matches, or there is nothing to
        prove it against. Never guesses in the parse's favour when a stated
        total exists and disagrees."""
        delta = self.commission_delta
        if delta is None:
            return True
        return abs(delta) <= CROSSCHECK_TOLERANCE

    @property
    def verifiable(self) -> bool:
        return self.stated_commission is not None

    def as_dict(self) -> dict[str, Any]:
        return {
            "parsed_total_premium": float(self.parsed_premium),
            "parsed_total_commission": float(self.parsed_commission),
            "stated_total_premium": (
                float(self.stated_premium) if self.stated_premium is not None else None
            ),
            "stated_total_commission": (
                float(self.stated_commission) if self.stated_commission is not None else None
            ),
            "crosscheck_ok": self.ok,
        }


def crosscheck(
    lines: list[dict[str, Any]],
    *,
    stated_premium: Any = None,
    stated_commission: Any = None,
) -> Crosscheck:
    result = Crosscheck(
        stated_premium=as_money(stated_premium),
        stated_commission=as_money(stated_commission),
    )
    for line in lines:
        result.parsed_commission += line.get("commission_amount") or Decimal("0")
        result.parsed_premium += line.get("gross_premium") or Decimal("0")
    return result


# --- staging -----------------------------------------------------------------

@dataclass
class StagedBatch:
    batch_id: str | None = None
    status: str = STATUS_PENDING_REVIEW
    filename: str = ""
    carrier: str | None = None
    line_count: int = 0
    crosscheck: Crosscheck = field(default_factory=Crosscheck)
    warnings: list[str] = field(default_factory=list)
    duplicate_of: str | None = None
    preview: dict[str, Any] = field(default_factory=dict)
    extraction_method: str = METHOD_CSV

    @property
    def is_ocr(self) -> bool:
        return self.extraction_method == METHOD_PDF_OCR

    @property
    def requires_confirmation(self) -> bool:
        """Does approving this batch need someone to say they checked the file?

        True for anything read out of a PDF. Not a judgement about this
        particular parse — it is a property of the format.
        """
        return requires_source_confirmation(self.extraction_method,
                                            is_ocr=self.is_ocr)

    @property
    def approvable(self) -> bool:
        """A batch may only be approved if it parsed, isn't a duplicate, and
        either matches the carrier's own totals or has none to match.

        A PDF batch is still *approvable* — it just cannot be approved silently.
        ``requires_confirmation`` is the extra thing the approver must supply.
        """
        return (
            self.status == STATUS_PENDING_REVIEW
            and self.line_count > 0
            and self.duplicate_of is None
            and self.crosscheck.ok
        )

    def as_dict(self) -> dict[str, Any]:
        return {
            "batch_id": self.batch_id,
            "status": self.status,
            "filename": self.filename,
            "carrier": self.carrier,
            "line_count": self.line_count,
            "approvable": self.approvable,
            "requires_confirmation": self.requires_confirmation,
            "extraction_method": self.extraction_method,
            "is_ocr": self.is_ocr,
            "duplicate_of": self.duplicate_of,
            "warnings": self.warnings,
            "crosscheck": {
                **self.crosscheck.as_dict(),
                "verifiable": self.crosscheck.verifiable,
                "commission_delta": (
                    float(self.crosscheck.commission_delta)
                    if self.crosscheck.commission_delta is not None else None
                ),
            },
            "preview": self.preview,
        }


def _match_preview(supa: "SupabaseClient", lines: list[dict[str, Any]]) -> dict[str, Any]:
    """Where these lines WOULD land, without writing anything.

    The reviewer is approving a set of consequences, not a file. Showing the
    unmatched policy numbers up front is the difference between a considered
    approval and a rubber stamp.
    """
    from hermes_core import book as ams_book
    from hermes.commissions.matching import (
        MATCH_CREATED, MATCH_EXACT, MATCH_NORMALIZED, MATCH_UNMATCHED,
        _index_book, _index_ledger, match_line,
    )

    try:
        ledger = supa.select("commission_ledger", columns="id,policy_number", limit=50000)
        exact_idx, norm_idx = _index_ledger(ledger)
    except Exception:  # noqa: BLE001
        log.exception("statement preview: ledger read failed")
        exact_idx, norm_idx = {}, {}
    try:
        book_idx = _index_book(ams_book.select_policies(
            supa, columns="policy_number,policy_guid,carrier,lines_of_business,"
                          "effective_date,expiration_date,premium_amount,annualized_premium",
            limit=50000))
    except Exception:  # noqa: BLE001
        log.exception("statement preview: book read failed")
        book_idx = {}

    counts = {MATCH_EXACT: 0, MATCH_NORMALIZED: 0, MATCH_CREATED: 0, MATCH_UNMATCHED: 0}
    unmatched: dict[str, int] = {}
    negatives = 0
    for line in lines:
        if (line.get("commission_amount") or Decimal("0")) < 0:
            negatives += 1
        result = match_line(line, ledger_by_exact=exact_idx,
                            ledger_by_normalized=norm_idx, book_by_normalized=book_idx)
        counts[result.kind] += 1
        if result.kind == MATCH_UNMATCHED:
            key = result.policy_number or "(blank)"
            unmatched[key] = unmatched.get(key, 0) + 1

    return {
        "will_link": counts[MATCH_EXACT] + counts[MATCH_NORMALIZED],
        "will_create_ledger_rows": counts[MATCH_CREATED],
        "will_be_unmatched": counts[MATCH_UNMATCHED],
        "unmatched_policy_numbers": unmatched,
        "negative_lines": negatives,
    }


def stage_statement(
    supa: "SupabaseClient",
    *,
    content: bytes,
    filename: str,
    uploaded_by: str,
    carrier: str | None = None,
    stated_premium: Any = None,
    stated_commission: Any = None,
) -> StagedBatch:
    """Parse and stage an uploaded statement. Writes NOTHING to the ledger."""
    digest = content_hash(content)

    existing = supa.select(BATCHES_TABLE, columns="id,ingest_status,source_file",
                           params={"content_hash": f"eq.{digest}"}, limit=1)
    if existing:
        prior = existing[0]
        return StagedBatch(
            batch_id=str(prior.get("id")), status=str(prior.get("ingest_status") or ""),
            filename=filename, carrier=carrier, duplicate_of=str(prior.get("id")),
            warnings=[f"identical file already uploaded as {prior.get('source_file')} "
                      f"(status {prior.get('ingest_status')})"],
        )

    parsed = parse_file(content, filename)
    lines, warnings, method = parsed.lines, parsed.warnings, parsed.method

    # A carrier parser reads the carrier's own totals off the statement. What the
    # uploader typed still wins — they are looking at the document — but when
    # they supply nothing, a crosscheck read from the file beats no crosscheck at
    # all, which is the case a bad parse walks straight through.
    if stated_premium in (None, "") and parsed.stated_premium is not None:
        stated_premium = parsed.stated_premium
    if stated_commission in (None, "") and parsed.stated_commission is not None:
        stated_commission = parsed.stated_commission
    carrier = carrier or parsed.carrier

    check = crosscheck(lines, stated_premium=stated_premium, stated_commission=stated_commission)

    status = STATUS_PENDING_REVIEW
    if not lines:
        status = STATUS_ERROR
        warnings.append("no statement lines could be parsed")
    elif not check.ok:
        warnings.append(
            f"parsed commission {check.parsed_commission} does not match the carrier's "
            f"stated {check.stated_commission} (delta {check.commission_delta}) — "
            "the parse is wrong; do not approve"
        )

    needs_confirmation = requires_source_confirmation(method)
    if needs_confirmation and lines:
        warnings.append(
            "read from a PDF — the column mapping is inferred, so every amount "
            "must be checked against the document before approval"
        )

    batch = supa.insert(BATCHES_TABLE, {
        "content_hash": digest,
        "source_file": filename,
        "carrier_name": carrier,
        "kind": "statement",
        "parser_key": parsed.parser_key,
        "extraction_method": method,
        "is_ocr": method == METHOD_PDF_OCR,
        "row_count": len(lines),
        "ingest_status": status,
        "uploaded_by": uploaded_by,
        "flags": {"warnings": warnings, "requires_confirmation": needs_confirmation},
        **check.as_dict(),
    })
    batch_id = str(batch.get("id"))

    for line in lines:
        payload = {
            "batch_id": batch_id,
            **{k: (float(v) if isinstance(v, Decimal) else v) for k, v in line.items()},
        }
        # carrier_name is NOT NULL on staging AND on commission_transactions, but
        # statement LINES rarely repeat the carrier — it is a property of the
        # statement, supplied on upload. Fall back to it rather than failing the
        # whole batch on a column the line was never going to carry.
        if not payload.get("carrier_name"):
            payload["carrier_name"] = carrier or "Unknown"
        supa.insert(STAGING_TABLE, payload)

    staged = StagedBatch(
        batch_id=batch_id, status=status, filename=filename,
        carrier=carrier, line_count=len(lines), crosscheck=check, warnings=warnings,
        extraction_method=method,
    )
    if lines:
        staged.preview = _match_preview(supa, lines)
    return staged


# --- commit ------------------------------------------------------------------

@dataclass
class CommitResult:
    batch_id: str
    statement_id: str | None = None
    committed: int = 0
    linked: int = 0
    created_ledger_rows: int = 0
    unmatched: int = 0
    rollup: str = ""
    errors: list[str] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        return {
            "batch_id": self.batch_id, "statement_id": self.statement_id,
            "committed": self.committed, "linked": self.linked,
            "created_ledger_rows": self.created_ledger_rows,
            "unmatched": self.unmatched, "rollup": self.rollup, "errors": self.errors,
        }


def commit_statement(
    supa: "SupabaseClient", *, batch_id: str, approved_by: str,
    confirmed_source: bool = False,
) -> CommitResult:
    """Promote a reviewed batch into the ledger. The approval gate is here.

    Refuses a batch that isn't pending review, parsed nothing, or failed its
    crosscheck — approving a bad parse is how fiction reaches a money surface.

    ``confirmed_source`` is the extra assertion a PDF batch needs: the approver
    states they compared the parsed lines against the document itself. It is a
    separate flag from the approval rather than part of it because approving a
    CSV and approving a picture of a statement are not the same act, and one
    button that means both would collapse them.
    """
    from hermes.commissions.matching import relink_unmatched
    from hermes.commissions.reconcile import run_rollup

    rows = supa.select(BATCHES_TABLE, columns="*", params={"id": f"eq.{batch_id}"}, limit=1)
    if not rows:
        raise ValueError(f"batch {batch_id} not found")
    batch = rows[0]

    status = str(batch.get("ingest_status") or "")
    if status != STATUS_PENDING_REVIEW:
        raise ValueError(f"batch is {status}, not {STATUS_PENDING_REVIEW}")
    if not batch.get("crosscheck_ok", True):
        raise ValueError(
            "crosscheck failed — parsed totals disagree with the carrier's stated "
            "totals; fix the parse rather than approving it"
        )

    method = str(batch.get("extraction_method") or "")
    if requires_source_confirmation(method, is_ocr=batch.get("is_ocr")) and not confirmed_source:
        raise ValueError(
            f"this batch was read from a PDF ({method}) — its columns are inferred, "
            "so it cannot be committed until the approver confirms the parsed "
            "lines match the document (confirmed_source)"
        )

    staged = supa.select(STAGING_TABLE, columns="*",
                         params={"batch_id": f"eq.{batch_id}"}, limit=50000)
    if not staged:
        raise ValueError("batch has no staged lines")

    result = CommitResult(batch_id=batch_id)

    statement = supa.insert(STATEMENTS_TABLE, {
        "carrier_name": batch.get("carrier_name"),
        "source_filename": batch.get("source_file"),
        "source_format": batch.get("extraction_method"),
        "carrier_stated_total_premium": batch.get("stated_total_premium"),
        "carrier_stated_total_commission": batch.get("stated_total_commission"),
        "row_count": len(staged),
        "upload_status": "parsed",
        "uploaded_by": approved_by,
    })
    result.statement_id = str(statement.get("id"))

    drop = {"id", "batch_id", "created_at"}
    for line in staged:
        payload = {k: v for k, v in line.items() if k not in drop}
        payload["statement_id"] = result.statement_id
        if not payload.get("carrier_name"):
            payload["carrier_name"] = batch.get("carrier_name") or "Unknown"
        # is_negative is GENERATED AS (commission_amount < 0) — Postgres derives
        # it from the amount we're already sending.
        payload = strip_generated(TRANSACTIONS_TABLE, payload)
        try:
            supa.insert(TRANSACTIONS_TABLE, payload)
            result.committed += 1
        except Exception as exc:  # noqa: BLE001 — one bad line must not lose the rest
            result.errors.append(f"line {line.get('policy_number')}: {exc}")

    # Attach the new lines, then recompute every touched ledger row.
    link = relink_unmatched(supa)
    result.linked = link.exact + link.normalized
    result.created_ledger_rows = link.ledger_rows_created
    result.unmatched = link.unmatched
    result.errors.extend(link.errors)

    result.rollup = run_rollup(supa).message

    # A batch where every line failed is not committed — marking it so would
    # hide the failure and block a retry behind the "already committed" guard.
    final_status = STATUS_COMMITTED if result.committed else STATUS_ERROR
    supa.update(BATCHES_TABLE, batch_id, {
        "ingest_status": final_status,
        "statement_id": result.statement_id,
        "reviewed_by": approved_by,
        "reviewed_at": datetime.now().astimezone().isoformat(),
    })
    log.info("statement %s committed by %s: %s lines", batch_id, approved_by, result.committed)
    return result


def reject_statement(
    supa: "SupabaseClient", *, batch_id: str, reviewed_by: str, reason: str | None = None,
) -> dict[str, Any]:
    """Reject a staged batch. Staging rows stay for diagnosis (FK is CASCADE, so
    they die with the batch only if the batch is deleted, which we never do)."""
    rows = supa.select(BATCHES_TABLE, columns="*", params={"id": f"eq.{batch_id}"}, limit=1)
    if not rows:
        raise ValueError(f"batch {batch_id} not found")
    flags = dict(rows[0].get("flags") or {})
    flags["rejected_reason"] = reason
    return supa.update(BATCHES_TABLE, batch_id, {
        "ingest_status": STATUS_REJECTED,
        "reviewed_by": reviewed_by,
        "reviewed_at": datetime.now().astimezone().isoformat(),
        "flags": flags,
    })
