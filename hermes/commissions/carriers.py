"""Carrier-specific statement parsers, and the sniff that picks one.

The generic reader in ``statements.py`` finds columns by alias, which is the
right default for a carrier nobody has mapped yet. But two carriers carry
knowledge no alias table can express, and it was previously encoded in the
browser (``src/parsers/*.ts``) — on the ingest path that wrote money straight to
Supabase with no staging and no approval. Retiring that path meant either
re-deriving this knowledge here, or losing it:

  PROGRESSIVE  Fixed column positions, and a second sheet holding the carrier's
               own stated totals — so the crosscheck that guards the commit can
               be read off the file instead of typed by hand. Its MVR chargeback
               lines carry money in *Agency Due* with zero commission; booked as
               commission they would read as a carrier paying nothing on a real
               policy.

  NEXT         An as-earned carrier. The statement prints both a cumulative
               "…Paid to Date" and an incremental "…Paid this Month". Only the
               incremental figure is summable across successive monthly
               statements; loading the cumulative one double-counts every month
               after the first. That single column choice is the whole reason
               this parser exists, and no alias table would ever guess it.

A carrier parser is tried first and the generic reader is the fallback. When
neither recognises the file the answer is "unmapped", never a guess — an
unrecognised statement is a mapping to write, not a parse to attempt.
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any

log = logging.getLogger(__name__)

PROGRESSIVE_V1 = "progressive_v1"
NEXT_V1 = "next_v1"

# NEXT statements carry no carrier column — the carrier is implied by the format.
# These are the names the ledger already uses; matching them is what lets a
# statement line find its policy.
PROGRESSIVE_CARRIER = "Progressive"
NEXT_CARRIER = "NEXT INS US CO"

# Excel serial dates: days since 1899-12-30. openpyxl usually hands back real
# datetimes, but a CSV exported from the same system can carry the raw serial.
_EXCEL_EPOCH = date(1899, 12, 30)
_EXCEL_SERIAL_MAX = 2958466        # 9999-12-31


@dataclass
class Sheet:
    """One tab of a workbook as a list of rows, header row included at index 0."""

    name: str
    rows: list[list[Any]] = field(default_factory=list)

    def row(self, index: int) -> list[Any]:
        return self.rows[index] if 0 <= index < len(self.rows) else []


@dataclass
class Workbook:
    sheets: list[Sheet] = field(default_factory=list)
    filename: str = ""

    def sheet(self, name: str) -> Sheet | None:
        for sheet in self.sheets:
            if sheet.name.strip().lower() == name.strip().lower():
                return sheet
        return None

    @property
    def first(self) -> Sheet | None:
        return self.sheets[0] if self.sheets else None


@dataclass
class CarrierParse:
    """What a carrier parser found: lines, plus what the carrier says it paid."""

    parser_key: str
    lines: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    carrier: str | None = None
    stated_premium: Decimal | None = None
    stated_commission: Decimal | None = None


# --- reading the file into rows ----------------------------------------------

def read_workbook(content: bytes, filename: str) -> Workbook:
    """Bytes -> sheets of raw cells. CSV becomes a single unnamed sheet."""
    suffix = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()
    if suffix in {"xlsx", "xlsm"}:
        return _read_xlsx(content, filename)
    if suffix in {"csv", "tsv", "txt"}:
        return _read_csv(content, filename)
    return Workbook(filename=filename)


def _read_xlsx(content: bytes, filename: str) -> Workbook:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - dependency is declared
        return Workbook(filename=filename)
    try:
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        log.info("carrier sniff: not a readable workbook: %s", filename)
        return Workbook(filename=filename)

    sheets: list[Sheet] = []
    for name in book.sheetnames:
        rows = [list(row) for row in book[name].iter_rows(values_only=True)]
        sheets.append(Sheet(name=name, rows=rows))
    book.close()
    return Workbook(sheets=sheets, filename=filename)


def _read_csv(content: bytes, filename: str) -> Workbook:
    import csv

    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows = [list(row) for row in csv.reader(io.StringIO(text), dialect=dialect)]
    return Workbook(sheets=[Sheet(name="", rows=rows)], filename=filename)


# --- cell coercion ------------------------------------------------------------

def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _label(value: Any) -> str:
    """Lowercase, whitespace-collapsed — how header cells are compared."""
    return re.sub(r"\s+", " ", _text(value)).lower()


def _cell(row: list[Any], index: int) -> Any:
    return row[index] if 0 <= index < len(row) else None


def as_iso_date(value: Any) -> str | None:
    """Statement date -> ISO, including the Excel serial form.

    ``statements.as_date`` handles the text and datetime cases; the serial case
    is specific to spreadsheet exports and lives here.
    """
    from hermes.commissions.statements import as_date

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        serial = int(round(float(value)))
        if 0 < serial < _EXCEL_SERIAL_MAX:
            return (_EXCEL_EPOCH + timedelta(days=serial)).isoformat()
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    parsed = as_date(value)
    return parsed.isoformat() if parsed else None


def _month_key(iso: str | None) -> int | None:
    if not iso:
        return None
    match = re.match(r"^(\d{4})-(\d{2})", iso)
    return int(match.group(1) + match.group(2)) if match else None


def _raw(row: list[Any]) -> dict[str, Any]:
    """Every cell as text, positional. The line as the carrier sent it."""
    return {str(i): (None if cell is None else str(cell)) for i, cell in enumerate(row)}


def _is_blank(row: list[Any]) -> bool:
    return not row or all(_text(cell) == "" for cell in row)


# --- Progressive --------------------------------------------------------------

# Fixed positions on the 'Detailed' sheet.
_P = {
    "insured": 0, "policy": 1, "eff_date": 2, "exp_date": 3, "prod": 4,
    "tran_code": 6, "tran_date": 7, "gross_premium": 8, "agency_due": 12,
    "rate": 13, "gross_comm": 14, "net_due": 15, "producer": 16,
    "agent_code": 17, "month_end": 18, "renewal_count": 19,
}


def _segment(prod: Any) -> str | None:
    """personal vs commercial, for the average-premium-by-segment rollup."""
    text = _text(prod)
    if not text:
        return None
    return "commercial" if re.search(r"commercial|comm\.?\s*veh", text, re.I) else "personal"


def _progressive_stated(summary: Sheet | None) -> tuple[Decimal | None, Decimal | None]:
    """The carrier's own totals, off the Summary sheet.

    Reading these is what makes the crosscheck automatic. Typed by hand they get
    skipped, and a batch with nothing to check against is a batch whose parse
    nobody can falsify.
    """
    from hermes.commissions.statements import as_money

    if summary is None:
        return None, None
    premium = commission = None
    for row in summary.rows:
        label = _label(_cell(row, 0))
        if label == "agent total":
            premium = as_money(_cell(row, 1))      # Net Written Premium — Current
            commission = as_money(_cell(row, 3))   # Commissions — Current
    return premium, commission


def parse_progressive(book: Workbook) -> CarrierParse:
    from hermes.commissions.statements import as_money, as_rate, normalize_transaction_type

    result = CarrierParse(parser_key=PROGRESSIVE_V1, carrier=PROGRESSIVE_CARRIER)
    detailed = book.sheet("Detailed")
    if detailed is None or len(detailed.rows) < 2:
        result.warnings.append("progressive: no 'Detailed' sheet with data rows")
        return result

    result.stated_premium, result.stated_commission = _progressive_stated(
        book.sheet("Summary")
    )
    if result.stated_commission is None:
        result.warnings.append(
            "progressive: no 'Agent Total' row on the Summary sheet — the parse "
            "has no carrier total to be checked against"
        )

    skipped = 0
    for row in detailed.rows[1:]:
        if _is_blank(row):
            skipped += 1
            continue

        policy_number = _text(_cell(row, _P["policy"]))
        gross_comm = as_money(_cell(row, _P["gross_comm"]))
        agency_due = as_money(_cell(row, _P["agency_due"]))
        if not policy_number and gross_comm is None:
            skipped += 1
            continue

        iso = as_iso_date(_cell(row, _P["tran_date"]))
        code = _text(_cell(row, _P["tran_code"]))

        # Money in Agency Due with no commission is an MVR chargeback, not a
        # commission line. Booking it as commission would read as the carrier
        # paying $0 on a live policy.
        is_fee = bool(agency_due) and agency_due != 0 and not gross_comm

        result.lines.append({
            "policy_number": policy_number,
            "insured_name": _text(_cell(row, _P["insured"])) or None,
            "carrier_name": PROGRESSIVE_CARRIER,
            "lob": _text(_cell(row, _P["prod"])) or None,
            "segment": _segment(_cell(row, _P["prod"])),
            "transaction_code": code or None,
            "transaction_type": "fee" if is_fee else normalize_transaction_type(code),
            "transaction_date": iso,
            "month_key": _month_key(iso) or _as_month_int(_cell(row, _P["month_end"])),
            "gross_premium": as_money(_cell(row, _P["gross_premium"])),
            "commission_rate": as_rate(_cell(row, _P["rate"])),
            "commission_amount": gross_comm,
            "fee_type": "MVR" if is_fee else None,
            "fee_amount": agency_due if is_fee else None,
            "raw_row": _raw(row),
        })

    if skipped:
        result.warnings.append(f"{skipped} row(s) skipped as blank or non-data")
    return result


def _as_month_int(value: Any) -> int | None:
    digits = re.sub(r"\D", "", _text(value))
    if len(digits) >= 6:
        candidate = int(digits[:6])
        if 190001 <= candidate <= 299912:
            return candidate
    return None


# --- NEXT ---------------------------------------------------------------------

_NEXT_COLS = {
    "policy": "policy number",
    "lob": "lob",
    "business": "business name",
    "statement_date": "statement date",
    "new_renewal": "new renewal",
    "agent_commission": "agent commission",
    "paid_this_month": "agency commission paid this month",
    "prem_this_month": "premium collected this month",
}


def _header_index(row: list[Any]) -> dict[str, int]:
    index: dict[str, int] = {}
    for position, cell in enumerate(row):
        label = _label(cell)
        if label:
            index.setdefault(label, position)
    return index


def looks_like_next(header: list[Any]) -> bool:
    """The distinctive trio. The incremental column is the giveaway."""
    index = _header_index(header)
    return all(
        _NEXT_COLS[key] in index
        for key in ("paid_this_month", "policy", "agent_commission")
    )


def parse_next(book: Workbook) -> CarrierParse:
    from hermes.commissions.statements import as_money, as_rate, normalize_transaction_type

    result = CarrierParse(parser_key=NEXT_V1, carrier=NEXT_CARRIER)
    sheet = book.first
    if sheet is None or len(sheet.rows) < 2:
        result.warnings.append("next: statement has no data rows")
        return result

    header = sheet.row(0)
    index = _header_index(header)
    names = [_text(cell) or f"col{i}" for i, cell in enumerate(header)]

    def value(row: list[Any], key: str) -> Any:
        position = index.get(_NEXT_COLS[key])
        return _cell(row, position) if position is not None else None

    skipped = 0
    for row in sheet.rows[1:]:
        if _is_blank(row):
            skipped += 1
            continue
        policy_number = _text(value(row, "policy"))
        if not policy_number:
            skipped += 1
            continue

        iso = as_iso_date(value(row, "statement_date"))
        code = _text(value(row, "new_renewal"))

        result.lines.append({
            "policy_number": policy_number,
            "insured_name": _text(value(row, "business")) or None,
            "carrier_name": NEXT_CARRIER,
            "lob": _text(value(row, "lob")) or None,
            # NEXT writes small-commercial only (GL/BP/PL/WC).
            "segment": "commercial",
            "transaction_code": code or None,
            "transaction_type": normalize_transaction_type(code),
            "transaction_date": iso,
            "month_key": _month_key(iso),
            # Incremental, not cumulative — see the module docstring.
            "gross_premium": as_money(value(row, "prem_this_month")),
            "commission_rate": as_rate(value(row, "agent_commission")),
            "commission_amount": as_money(value(row, "paid_this_month")),
            "fee_type": None,
            "fee_amount": None,
            # Named cells here: NEXT's cumulative columns are the audit trail for
            # why the incremental one was chosen, and they are worth reading back.
            "raw_row": {
                names[i]: (None if _cell(row, i) is None else str(_cell(row, i)))
                for i in range(len(names))
            },
        })

    if skipped:
        result.warnings.append(f"{skipped} row(s) skipped as blank or without a policy number")
    result.warnings.append(
        "NEXT is as-earned: this loads the INCREMENTAL 'paid this month' column, "
        "so successive monthly statements sum to the term total"
    )
    return result


# --- the sniff ----------------------------------------------------------------

def detect(book: Workbook) -> str | None:
    """Which carrier parser handles this file? ``None`` means unmapped.

    Filename first for Progressive because its export is named distinctively,
    then a content sniff for both — a renamed file must still be recognised, and
    a lookalike must still be refused.
    """
    if re.search(r"detailedstatement", book.filename or "", re.I):
        if book.sheet("Detailed") is not None:
            return PROGRESSIVE_V1

    detailed = book.sheet("Detailed")
    if detailed is not None:
        header = detailed.row(0)
        if (_text(_cell(header, _P["tran_code"])) == "Tran Code"
                and _text(_cell(header, _P["policy"])) == "Policy Number"):
            return PROGRESSIVE_V1

    first = book.first
    if first is not None and first.rows and looks_like_next(first.row(0)):
        return NEXT_V1

    return None


_PARSERS = {PROGRESSIVE_V1: parse_progressive, NEXT_V1: parse_next}


def parse_carrier(content: bytes, filename: str) -> CarrierParse | None:
    """Parse with a carrier-specific parser, or ``None`` if none recognises it."""
    book = read_workbook(content, filename)
    if not book.sheets:
        return None
    key = detect(book)
    if key is None:
        return None
    try:
        return _PARSERS[key](book)
    except Exception:  # noqa: BLE001 — a broken carrier parser falls back to generic
        log.exception("carrier parser %s failed on %s", key, filename)
        return None
